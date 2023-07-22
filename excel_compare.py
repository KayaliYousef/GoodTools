import openpyxl
from openpyxl.styles import PatternFill


def compare_excel(path_to_first_excel, path_to_second_excel):
    # Open the first Excel file
    wb1 = openpyxl.load_workbook(path_to_first_excel)
    ws1 = wb1.active
    wb2 = openpyxl.load_workbook(path_to_second_excel)
    ws2 = wb2.active
    # Set red fill color
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Iterate over the cells in the first file and fill the background color with red if different
    for row in range(1, ws1.max_row + 1):
        for col in range(1, ws1.max_column + 1):
            file1_cell = ws1.cell(row=row, column=col)
            file2_cell = ws2.cell(row=row, column=col)
            if file1_cell.value != file2_cell.value:
                file1_cell.fill = red_fill
                file2_cell.fill = red_fill

    # Save the modified Excel file
    wb1.save('modified' + path_to_first_excel)
    wb2.save('modified' + path_to_second_excel)
